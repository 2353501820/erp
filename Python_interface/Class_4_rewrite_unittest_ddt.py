# -*- coding: utf-8 -*-
"""
--------------------------------------------------
# file : Class_4_rewrite_unittest_ddt.py
# author : liushen
# time : 2020/11/23 0023 21:33
---------------------------------------------------
"""
# ddt，是一种数据驱动思想，是辅助加载测试用例的一个库
# 如果使用for循环，实例方法执行的时候如果抛出异常，那么这个实例方法将会终止执行，如果没有终止有for循环的话，代码会一直执行，直到抛出异常
import unittest
from ddt import ddt,data,unpack
import inspect
from openpyxl import load_workbook
from collections import namedtuple
from pythonbase_class_1.Class_9_1_MathOperation import MathOperation

test_case = (10,0,True,False,None,2,12.1,"","静宝")
@ddt
class Test01(unittest.TestCase):   # ddt是用来装饰类的，需要与data装饰器一起使用，装饰器就是给类或者函数添加额外的功能
	"""
    使用ddt来加载测试数据
	"""
	# @data(*test_case)  # 将序列类型拆包为多个位置参数
	@data(10,0,True,False,None,2,12.1,"","静宝")
	def test_1(self,val):   # 创建测试用例
		print("\nrunning test method:{}".format(inspect.stack()[0][3]))
		# self.assertTrue(val)
		print("值为{}，类型为{}".format(val,type(val)))
		self.assertTrue(val)


# if __name__ == '__main__':
# 	unittest.main()

# 执行了多条用例，用例执行的条数与data装饰器的（位置参数）个数一致
# 每执行一条用例会自动将一个参数传给val，当最后一个参数传给val，且用例执行结束过后，程序执行完毕
# 这个过程相当于遍历参数
# ddt和data要一起使用


# 使用@unpack装饰器来进一步拆包(用的不多，了解即可)

@ddt
class Test02(unittest.TestCase):   # ddt是用来装饰类的，需要与data装饰器一起使用，装饰器就是给类或者函数添加额外的功能
	"""
    使用ddt来加载测试数据
	"""
	@data([True,False,None],("二姨","静宝",0))
	@unpack   #使用unpack，那么所有的元素要能支持拆包，序列类型可以拆包（list,dict,tuple）
	def test_2(self,*val):   # 创建测试用例
		print("\nrunning test method:{}".format(inspect.stack()[0][3]))
		# self.assertTrue(val)
		print("值为{}，类型为{}".format(val,type(val)))
		self.assertTrue(val)


# if __name__ == '__main__':
# 	unittest.main()


# 使用ddt来重写之前的测试用例


wb = load_workbook("test.xlsx")
ws = wb.active
sheet_head_tuple = tuple(ws.iter_rows(max_row=1,values_only=True))[0]
test_list = []
test = namedtuple("test",sheet_head_tuple)
for test_data in ws.iter_rows(min_row=2,max_row=5,values_only=True):
	test_list.append(test(*test_data))   # for循环尽量不要与关键字重名


@ddt
class TestMulti(unittest.TestCase):  # unittest规则：每一条用例，使用实例方法来测试，并且实例方法吗要以test_开头
	"""
	测试两数相乘
	"""
	@classmethod
	def tearDownClass(cls):
		wb.save("test.xlsx")


	@data(*test_list)
	def test_negative_multi(self,data_namedtuple):
		"""
		1.两个负数相乘
		:return:
		"""
		print("\nrunning test method:{}".format(inspect.stack()[0][3]))
		# 		获取两个负数相乘的结果
		case_id = data_namedtuple.case_id
		msg = data_namedtuple.title
		l_data = data_namedtuple.l_data
		r_data = data_namedtuple.r_data
		except_result = data_namedtuple.expected
		real_result = MathOperation(l_data, r_data).multiply()  # 类+() 是对象 用对象调用类中的实例multiply()
		# 		验证实际结果与预期是否一致
		ws.cell(row=case_id+1,column=6,value=real_result)   # 将实际结果写入Excel
		# except_result = 4   # 不需要手动添加期待值
		# 第一种方法使用条件判断不能实现功能
		# if real_result == except_result:
		# 	print("pass")
		# else:
		# 	print("fail")
		try:
			self.assertEqual(except_result, real_result, msg="测试{}失败".format(msg))
		except AssertionError as e:
			print("这里需要使用日志器来记录日志")
			print("具体异常为：{}".format(e))
			ws.cell(row=case_id+1,column=7,value="Fail")
			self.assertRaises(TypeError)   # assertRaises可以直接断言异常
			# raise关键字是将某个异常主动抛出
			raise e
		else:
			ws.cell(row=case_id+1,column=7,value="Pass")


if __name__ == '__main__':
	unittest.main()
