# -*- coding: utf-8 -*-
"""
--------------------------------------------------
# file : Class_1_excel_handle.py
# author : liushen
# time : 2020/11/13 0013 19:22
---------------------------------------------------
"""
# 将测试数据使用Excel来处理
# Python中处理Excel的模块非常多，比如xlrd(只能读Excel)、xlwt（只能写Excel），读写分开不太方便
# 使用openpyxl，同时支持读写操作，需要安装之后才能使用
# pip install openpyxl
# 如果报错，需要将pip更新一下，python -m pip install -upgrade pip
# 注意：openpyxl只能处理后缀为xlsx的Excel文件，其他后缀的文件不能处理
from openpyxl import load_workbook   # 可以对已存在的Excel进行读写操作
from openpyxl import workbook   # 可以新建Excel文件


# 使用load_workbook来实现Excel读写
# 1.打开Excel文件（已存在）
wb = load_workbook("test.xlsx")
# 一个Excel文件由哪些部分组成？
# 文件名，表单，单元格
# 第一个参数名为文件名，会返回一个workbook对象（相当于整个Excel文件）

# 2.定位表单
# 方法一：
# ws = wb["sheet3"]   # ws为worksheet对象（相当于Excel中的一个表单）

# 方法二：
ws = wb.active   # 获取第一个表单

# 3.定位单元格cell
# 方法一：
# one_cell = ws.cell(row=2,column=2)   # one_cell为cell对象（相当于一个表单中的某个单元格）
# print(one_cell.value)   # 获取单元格中的值

# 方法二：每遍历一次处理一个单元格
# worksheet对象中有如下重要的属性：
# max_row：单元格最大行
# min_row：单元格最小行
# max_column：单元格最大列
# min_column：单元格最小列

# worksheet对象中有如下重要的方法：
# iter_rows：返回一个生成器，是由将每一行的数据作为元组组成的
# iter_columns：返回一个生成器，是由将每一列的数据作为元组组成的
# for row_index in range(ws.min_row+1,ws.max_row+1):   # 第一行不取，所以索引为2，最后一行取不到，所以最后一行+1
# 	for column_index in range(ws.min_column,ws.max_column+1):
# 		data = ws.cell(row_index,column_index)
# 		print("值{},类型{}".format(data,type(data)))

# 如果获取的是数字（int、float），那么读取的数字也是数字类型
# 所有的非数字类型，读取的都是字符串类型
# 如果值为空，值为None

# 方法三：
# for row_tuple in ws.iter_rows(min_row=2):   # 每遍历一次，会将某一行的所有单元格对象（cell对象）组成一个元组返回，max_row,min_column,max_column都是默认值，可以不用指定
# 	# print(row_tuple)
# 	for one_cell in row_tuple:
# 		data = one_cell.value
# 		print("值{},类型{}".format(data, type(data)))


# for row_tuple in ws.iter_rows(min_row=2,values_only=True):
# 	# values_only=True 每遍历一次，会将某一行的所有单元格（cell对象）的值组成一个元组返回
# 	for data in row_tuple:
# 		print("值{},类型{}".format(data, type(data)))

# 读数据不需要关闭，也不需要保存excel
# 方法四：指定需要处理的所有单元格
sheets = ws["A2:G6"]   # 返回的是嵌套cell对象元组的元组，内层元组是由每一行的每一个单元格的cell对象组成
for row in sheets:
	for cell_tuple in row:
		data = cell_tuple.value
		print("值{},类型{}".format(data, type(data)))
# 这里返回的是元组，而方法三中返回的是生成器，从系统优化的角度讲，方法三要优于方法四

# 从Excel中的数字类型读取到Python中之后也为数字类型（int、float、bool等）
# 所有的非数字类型都会被当做字符串
# 如何将Excel中的列表转换为list列表类型？
# 可以使用eval函数来处理（最好的方法）
# 使用字符串切割的方法（非常麻烦）
# 可以使用正则表达式来处理（也是可以的，也非常通用，还是有些麻烦）

# b6_value = ws.cell(6,2).value
# b6_list = eval(b6_value)
# print(b6_list,type(b6_list))

# 在Excel中保存Python中的数据类型时，一定要满足他的语法规范 如果写成none，一定要打冒号认定为字符串
# eval("a")第一个参数必须是字符串，且符合Python语法规范；可以是python中的数据类型，也可以是表达式（算术运算、逻辑运算等）

# 尽量不要这样用（input）,用户可能会传些非法的代码（病毒），python执行之后会有很大的危害
# one_str = input("请输入：")
# print(eval(one_str))

# （伪代码）对异常进行捕获
# if isinstance(one_data(int,float,bool)):
# 	print("不需要使用eval来转换")
# else:
# 	try:
# 		eval(one_data)
# 	except Exception as e:
# 		print(e)

pass