# -*- coding: utf-8 -*-
"""
--------------------------------------------------
# file : Class_2_excel_handle_writing.py
# author : liushen
# time : 2020/11/18 0018 19:11
---------------------------------------------------
"""
# 如何将执行的测试用例写入Excel当中
from openpyxl import load_workbook

# 1.打开Excel文件
wb = load_workbook("test.xlsx")

# 2.定位sheet
# ws = wb["Sheet3"]
ws = wb.active
# 3.定位单元格
# 方法一：
# ws.cell(2,6).value = 8
#
# # 方法二：
# ws.cell(2,7,value="Pass")
#
# # 方法三：添加一行的内容
# row_data = ((6,"负数与正数相乘",10,-3,-30,None,None),(7,"正数与零相乘",0,7,0,None,None))
# # 可以使用append来添加一行的数据
# # ws.append(row_data[0])   # 可以给他传一个任意一个序列类型
# # 如果想要批量添加可以使用for循环
# for row in row_data:
# 	ws.append(row)

# 4.保存Excel
# wb.save("test.xlsx")   # 保存的文件名要与打开的Excel文件名一致，只要修改了Excel就一定要保存

# 报“Permission denied: 'test.xlsx'”错误说明文件处在打开状态，写入文件时一定要保证Excel处于关闭状态



# 将数据从Excel读取出来，在python中来处理
# 获取表头的信息
# ws.iter_rows(max_row=1)
# tuple(ws.iter_rows(max_row=1)) 和 ws[A1:G1]一样返回元组的元组
sheet_head_tuple = tuple(ws.iter_rows(max_row=1,values_only=True))[0]   # 返回以字符串组成的元组
#
# # 方法一：
# # 生成器需要for循环遍历触发才能获取值
# # 将获取的数据转化为字典
# # 定义一个嵌套字典的列表来保存用例
# test_list = []
# for data in ws.iter_rows(min_row=2,max_row=5,values_only=True):
# 	test_list.append({
# 		"case_id":data[0],
# 		"title": data[1],
# 		"l_data":data[2],
# 		"r_data":data[3],
# 		"expected":data[4],
# 		"actual":data[5],
# 		"result":data[6]
# 	})
# 表头信息若是手写输入特别麻烦，拓展性不好

# 方法二：
# zip 相当于一一对应，将第一个序列类型与第二个序列类型中的元素进行一一对应然后组合为一个元组
# 返回一个zip对象的生成器、map，reduce(快弃用)
# one_tuple = ("静宝",3,"快上幼儿园的小不点")
# one_list = ["二姨",26,"静宝的二姨"]
# print(dict(zip(one_tuple,one_list)))
# print(list(zip(one_tuple,one_list)))
# print(tuple(zip(one_tuple,one_list)))
test_list = []
# for data in ws.iter_rows(min_row=2,max_row=5,values_only=True):
# 	test_list.append(dict(zip(sheet_head_tuple,data)))

# 方法三：（namedtuple）
from collections import namedtuple
test = namedtuple("test",sheet_head_tuple)   # 创建一个元组类
for data in ws.iter_rows(min_row=2,max_row=5,values_only=True):   # 每次遍历，返回由某行所有单元格值
	test_list.append(test(*data))
pass