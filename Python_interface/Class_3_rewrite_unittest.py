# -*- coding: utf-8 -*-
"""
--------------------------------------------------
# file : Class_3_rewrite_unittest.py
# author : liushen
# time : 2020/11/21 0021 19:32
---------------------------------------------------
"""
# 重写之前的单元测试
import unittest
import inspect


from openpyxl import load_workbook
from collections import namedtuple

# 导入自定义模块
from pythonbase_class_1.Class_9_1_MathOperation import MathOperation


# 1.打开Excel文件
wb = load_workbook("test.xlsx")

# 2.定位sheet
# ws = wb["Sheet3"]
ws = wb.active

# 3.定位单元格
sheet_head_tuple = tuple(ws.iter_rows(max_row=1,values_only=True))[0]   # 返回以字符串组成的元组
test_list = []
test = namedtuple("test",sheet_head_tuple)   # 创建一个元组类
for data in ws.iter_rows(min_row=2,max_row=5,values_only=True):   # 每次遍历，返回由某行所有单元格值
	test_list.append(test(*data))


class TestMulti(unittest.TestCase):  # unittest规则：每一条用例，使用实例方法来测试，并且实例方法吗要以test_开头
	"""
	测试两数相乘
	"""
	@classmethod
	def tearDownClass(cls):
		wb.save("test.xlsx")


	def test_negative_multi(self):
		"""
		1.两个负数相乘
		:return:
		"""
		print("\nrunning test method:{}".format(inspect.stack()[0][3]))
		# case_logger.info("\nrunning test method:{}".format(inspect.stack()[0][3]))   #记录日志
		# 		获取两个负数相乘的结果
		data_namedtuple = test_list.pop(0)
		case_id = data_namedtuple.case_id
		msg = data_namedtuple.title
		l_data = data_namedtuple.l_data
		r_data = data_namedtuple.r_data
		except_result = data_namedtuple.expected
		real_result = MathOperation(l_data, r_data).multiply()  # 类+() 是对象 用对象调用类中的实例multiply()
		# 		验证实际结果与预期是否一致
		ws.cell(row=case_id+1,column=6,value=real_result)   # 将实际结果写入Excel
		# except_result = 4   # 不需要手动添加期待值
		# 第一种方法使用条件判断不能实现功能
		# if real_result == except_result:
		# 	print("pass")
		# else:
		# 	print("fail")
		try:
			self.assertEqual(except_result, real_result, msg="测试{}失败".format(msg))
		except AssertionError as e:
			print("这里需要使用日志器来记录日志")
			# case_logging.error("具体异常为：{}".format(e))
			print("具体异常为：{}".format(e))
			ws.cell(row=case_id+1,column=7,value="Fail")
			# raise关键字是将某个异常主动抛出
			raise e
		else:
			ws.cell(row=case_id+1,column=7,value="Pass")

	def test_two_zero_multi(self):
		"""
		2.两个零相乘
		:return:
		"""
		print("\nrunning test method:{}".format(inspect.stack()[0][3]))
		data_namedtuple = test_list.pop(0)
		case_id = data_namedtuple.case_id
		msg = data_namedtuple.title
		l_data = data_namedtuple.l_data
		r_data = data_namedtuple.r_data
		except_result = data_namedtuple.expected
		real_result = MathOperation(l_data, r_data).multiply()  # 类+() 是对象 用对象调用类中的实例multiply()
		# 		验证实际结果与预期是否一致
		ws.cell(row=case_id + 1, column=6, value=real_result)  # 将实际结果写入Excel
		# except_result = 4   # 不需要手动添加期待值
		# 第一种方法使用条件判断不能实现功能
		# if real_result == except_result:
		# 	print("pass")
		# else:
		# 	print("fail")
		try:
			self.assertEqual(except_result, real_result, msg="测试{}失败".format(msg))
		except AssertionError as e:
			print("这里需要使用日志器来记录日志")
			print("具体异常为：{}".format(e))
			ws.cell(row=case_id + 1, column=7, value="Fail")
			# raise关键字是将某个异常主动抛出
			raise e
		else:
			ws.cell(row=case_id + 1, column=7, value="Pass")

	def test_two_pos_multi(self):
		"""
		3.两个正数相乘
		:return:
		"""
		print("\nrunning test method:{}".format(inspect.stack()[0][3]))
		data_namedtuple = test_list.pop(0)
		case_id = data_namedtuple.case_id
		msg = data_namedtuple.title
		l_data = data_namedtuple.l_data
		r_data = data_namedtuple.r_data
		except_result = data_namedtuple.expected
		real_result = MathOperation(l_data, r_data).multiply()  # 类+() 是对象 用对象调用类中的实例multiply()
		# 		验证实际结果与预期是否一致
		ws.cell(row=case_id + 1, column=6, value=real_result)  # 将实际结果写入Excel
		# except_result = 4   # 不需要手动添加期待值
		# 第一种方法使用条件判断不能实现功能
		# if real_result == except_result:
		# 	print("pass")
		# else:
		# 	print("fail")
		try:
			self.assertEqual(except_result, real_result, msg="测试{}失败".format(msg))
		except AssertionError as e:
			print("这里需要使用日志器来记录日志")
			print("具体异常为：{}".format(e))
			ws.cell(row=case_id + 1, column=7, value="Fail")
			# raise关键字是将某个异常主动抛出
			raise e
		else:
			ws.cell(row=case_id + 1, column=7, value="Pass")

	def test_neg_pos_multi(self):
		"""
		4.一个正数和一个负数相乘
		:return:
		"""
		data_namedtuple = test_list.pop(0)
		case_id = data_namedtuple.case_id
		msg = data_namedtuple.title
		l_data = data_namedtuple.l_data
		r_data = data_namedtuple.r_data
		except_result = data_namedtuple.expected
		real_result = MathOperation(l_data, r_data).multiply()  # 类+() 是对象 用对象调用类中的实例multiply()
		# 		验证实际结果与预期是否一致
		ws.cell(row=case_id + 1, column=6, value=real_result)  # 将实际结果写入Excel
		# except_result = 4   # 不需要手动添加期待值
		# 第一种方法使用条件判断不能实现功能
		# if real_result == except_result:
		# 	print("pass")
		# else:
		# 	print("fail")
		try:
			self.assertEqual(except_result, real_result, msg="测试{}失败".format(msg))
		except AssertionError as e:
			print("这里需要使用日志器来记录日志")
			print("具体异常为：{}".format(e))
			ws.cell(row=case_id + 1, column=7, value="Fail")
			# raise关键字是将某个异常主动抛出
			raise e
		else:
			ws.cell(row=case_id + 1, column=7, value="Pass")


if __name__ == '__main__':
	unittest.main()

# 以上方法每条用例代码完全一致，冗余严重；
# 这种设计中Excel的用例数量与单元测试实例方法数一致，若不一致会报错，程序拓展性极差
# 对Excel进行读写操作，可读性非常差，程序封装性不好，程序不够灵活，代码很难复用

