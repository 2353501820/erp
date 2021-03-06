# -*- coding: utf-8 -*-
"""
--------------------------------------------------
# file : Class_14_excel_add_unittest_Html.py
# author : liushen
# time : 2020/12/16 0016 18:21
---------------------------------------------------
"""

import unittest
# 修改ddt源码，把class13中的很长的字符拼接test_negative_multi_2_test_case_id_2__title__负数与正数相乘___l_data__3__r_data_4__expected_12__actual__12__result__Fail__改短
from Python_interface.ddt import ddt,data
import inspect
from collections import namedtuple
from openpyxl import load_workbook

from pythonbase_class_1.Class_9_1_MathOperation import MathOperation
# from Python_interface.Class_5_excel_encapsulation import HandleExcel
# from Python_interface.Class_8_ParserConfigFile__call__ import Handleconfig
# from Python_interface.Class_11_Rewrite_excel_encapsulation import do_excel
from Python_interface.Class_11_Rewrite_excel_encapsulation import HandleExcel   # 因为运算加法在第二个sheet里，而do_excel对象指定的是第一个sheet，所有我们直接导类，不导对象
from Python_interface.Class_8_ParserConfigFile__call__ import do_config
from Python_interface.Class_10_Log_encapsulation import do_log

do_excel = HandleExcel(do_config("file path","test_path"),"add")   # 重新定义do_excel

@ddt
class TestAdd(unittest.TestCase):  # unittest规则：每一条用例，使用实例方法来测试，并且实例方法吗要以test_开头
	"""
	测试两数相加
	"""
	# config = Handleconfig()
	# # test_obj = HandleExcel("test.xlsx")
	# test_obj = HandleExcel(config("file path","test_path"))
	# tests = test_obj.get_tests()
	tests = do_excel.get_tests()

	@classmethod
	def setUpClass(cls):
		"""
		重写父类类方法，在实例方法执行之前会被调用一次
		:return:
		"""
		# print("\n{:=^40s}".format("开始执行用例"))
		do_log.info("\n{:=^40s}".format("开始执行用例"))
		# cls.file_name = cls.config("file_path","log_path")
		# print("打开【{}】文件".format(cls.file_name))
		# cls.file = open(cls.file_name,mode="a",encoding="utf-8")
		# cls.file.write("\n{:=^40s}\n".format("开始执行用例"))

	@classmethod
	def tearDownClass(cls):
		"""
		重写父类类方法，在实例方法执行结束之后会被调用一次
		:return:
		"""
		# print("\n{:=^40s}".format("用例执行结束"))
		do_log.info("\n{:=^40s}".format("用例执行结束"))
		# cls.file.write("\n{:=^40s}\n".format("用例执行结束"))
		# print("关闭【{}】文件".format(cls.file_name))
		# cls.file.close()

	@data(*tests)
	def test_negative_multi(self,data_namedtuple):
		"""
		1.两个负数相加
		:return:
		"""
		# print("\nrunning test method:{}".format(inspect.stack()[0][3]))
		do_log.info("\nrunning test method:{}".format(inspect.stack()[0][3]))
		# 		获取两个负数相乘的结果
		case_id = data_namedtuple.case_id
		msg = data_namedtuple.title
		l_data = data_namedtuple.l_data
		r_data = data_namedtuple.r_data
		except_result = data_namedtuple.expected
		# real_result = MathOperation(l_data, r_data).multiply()  # 类+() 是对象 用对象调用类中的实例multiply()
		real_result = MathOperation(l_data, r_data).add()  # 类+() 是对象 用对象调用类中的实例add()
		# 将实际结果写入Excel
		# run_success_msg = TestMulti("msg","success_result")
		run_success_msg = do_config("msg", "success_result")
		# run_fail_msg = TestMulti("msg", "fail_result")
		run_fail_msg = do_config("msg", "fail_result")
		try:
			self.assertEqual(except_result, real_result, msg="测试{}失败".format(msg))
		except AssertionError as e:
			# print("这里需要使用日志器来记录日志")
			# print("具体异常为：{}".format(e))
			do_log.error("具体异常为：{}".format(e))
			# ws.cell(row=case_id+1,column=7,value="Fail")
			# self.test_obj.write_result(row=case_id+1,actual=real_result,result="Fail")
			# self.test_obj.write_result(row=case_id+1,actual=real_result,result=self.config("msg","fail_result"))
			do_excel.write_result(row=case_id + 1, actual=real_result, result=do_config("msg", "fail_result"))
			# self.assertRaises(TypeError)   # assertRaises可以直接断言异常
			# raise关键字是将某个异常主动抛出
			raise e
		else:
			# ws.cell(row=case_id+1,column=7,value="Pass")
			# self.test_obj.write_result(row=case_id+1,actual=real_result,result="Pass")
			do_excel.write_result(row=case_id+1,actual=real_result,result=do_config("msg","success_result"))
			# self.file.write("{}，执行结果为：{}\n".format(msg,run_success_msg))

if __name__ == '__main__':
	unittest.main()
