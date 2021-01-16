# -*- coding: utf-8 -*-
"""
--------------------------------------------------
# file : Class_5_excel_encapsulation.py
# author : gaobo
# time : 2020/11/25 0025 18:45
---------------------------------------------------
"""
from openpyxl import load_workbook
from collections import namedtuple
from ddt import ddt,data
import unittest
import inspect
from pythonbase_class_1.Class_9_1_MathOperation import MathOperation
from Python_interface.Class_8_ParserConfigFile__call__ import Handleconfig


class HandleExcel(object):   # 如果你希望其他的实例方法可以调用，那么就在实例方法中使用self
	"""
	定义处理Excel的类
	"""
	config = Handleconfig()
	def __init__(self,filename,sheetname=None):
		self.filename = filename
		self.sheetname = sheetname
		self.wb = load_workbook(self.filename)
		# self.ws = self.wb[self.sheetname] if self.sheetname is not None else self.wb.active   # 三元运算，为真执行左边，为假执行else右边数据
		if self.sheetname is None:   #如果没有穿sheetname这个参数，那么默认获取第一个表单
			self.ws = self.wb.active
		else:
			self.ws = self.wb[self.sheetname]
		self.sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1,values_only=True))[0]
		self.test = namedtuple("test",self.sheet_head_tuple)
		self.test_list = []   # 定义一个存放所有test命名元组对象的空列表

	def get_tests(self):
		"""
		获取所有的测试用例
		:return:存放test命名元组的列表
		"""
		for data in self.ws.iter_rows(min_row=2,max_row=5,values_only=True):
			self.test_list.append(self.test(*data))
		return self.test_list

	def get_row(self,row):
		"""
		获取某一条测试用例
		:param row:行号
		:return:一个test对象
		"""
		if isinstance(row,int) and (2 <= row <= self.ws.max_row):
			return tuple(self.ws.iter_rows(min_row=row,max_row=row,values_only=True))[0]
		else:
			print("传入的不是大于1的整数")

	def write_result(self,row,actual,result):
		"""
		将实际值与测试用例执行的结果保存到Excel中
		:param row:保存到哪一行
		:param actual:实际结果
		:param result:测试用例执行结果'Pass''Fail'
		:return:
		"""
		if isinstance(row,int) and (2 <= row <= self.ws.max_row):
			# self.ws.cell(row=row,column=6,value=actual)
			self.ws.cell(row=row,column=self.config("excel","actual_col"),value=actual)
			# self.ws.cell(row=row,column=7,value=result)
			self.ws.cell(row=row,column=HandleExcel.config("excel","result_col"),value=result)
			self.wb.save(self.filename)


if __name__ == '__main__':
	file_name = "test.xlsx"
	one_excel = HandleExcel(filename=file_name)
	test = one_excel.get_tests()
	print(test)



# 总结
# 如何来封装一些操作
# 1.明确需求（读取所有的测试用例，读取一条用例，写入结果）
# 2.将一些写死的数据把它提取出来（例如文件名），往往我们把它当做属性
# 3.如果其他实例方法中需要调用构造方法中的变量那么需要将其定义为属性